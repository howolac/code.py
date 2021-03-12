#!python

import os, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

os.chdir('c:\\users\\howolac\\desktop')

wb = openpyxl.load_workbook('1.xlsx')
sheet1 = wb.get_sheet_by_name('Sheet1')

#print(tuple(sheet1['A1':'C7']))
#遍历所有行
for j in range(sheet1.max_row):
    for i in range(sheet1.max_column):
        print(get_column_letter(i + 1) + ' ' + str(i + 1), end='')
        print(sheet1.cell(row=j + 1, column=i + 1).value)
    print('---------------end of row---------------')
#每一在结尾打印出结束信息
